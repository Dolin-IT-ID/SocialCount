import pandas as pd
import io
from datetime import datetime
from typing import List, Dict
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from i18n import get_text

class DataExporter:
    def __init__(self):
        pass
    
    def prepare_data_for_export(self, video_data: List[Dict]) -> pd.DataFrame:
        """Menyiapkan data untuk ekspor dengan format yang rapi"""
        if not video_data:
            return pd.DataFrame()
        
        # Konversi data MongoDB ke format yang lebih mudah dibaca
        export_data = []
        for video in video_data:
            row = {
                get_text('export_id'): str(video.get('_id', '')),
                get_text('export_creator_name'): video.get('nama_pembuat', ''),
                get_text('export_account_name'): video.get('nama_akun', ''),
                get_text('export_video_url'): video.get('url_video', ''),
                get_text('export_platform'): video.get('platform', ''),
                get_text('export_video_title'): video.get('title', ''),
                get_text('export_description'): video.get('description', '')[:100] + '...' if video.get('description', '') else '',
                get_text('export_views'): video.get('views', 0),
                get_text('export_likes'): video.get('likes', 0),
                get_text('export_comments'): video.get('comments', 0),
                get_text('export_shares'): video.get('shares', 0),
                get_text('export_duration'): video.get('duration', ''),
                get_text('export_upload_date'): video.get('upload_date', ''),
                get_text('export_calculation_time'): video.get('waktu_penghitungan', ''),
                get_text('export_created_at'): video.get('created_at', ''),
                get_text('export_updated_at'): video.get('updated_at', '')
            }
            export_data.append(row)
        
        df = pd.DataFrame(export_data)
        
        # Format tanggal untuk tampilan yang lebih baik
        date_columns = ['Waktu Penghitungan', 'Created At', 'Updated At']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')
        
        return df
    
    def export_to_csv(self, video_data: List[Dict]) -> bytes:
        """Ekspor data ke format CSV"""
        df = self.prepare_data_for_export(video_data)
        if df.empty:
            return b''
        
        # Konversi ke CSV dengan encoding UTF-8
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8')
        return csv_buffer.getvalue().encode('utf-8')
    
    def export_to_excel(self, video_data: List[Dict]) -> bytes:
        """Ekspor data ke format Excel dengan styling"""
        df = self.prepare_data_for_export(video_data)
        if df.empty:
            return b''
        
        # Buat workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Video Summary"
        
        # Styling untuk header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Tambahkan data ke worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply styling ke header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Simpan ke buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def create_download_link(self, data: bytes, filename: str, file_format: str) -> None:
        """Membuat link download untuk file"""
        if file_format.lower() == 'csv':
            mime_type = 'text/csv'
        elif file_format.lower() == 'excel':
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mime_type = 'application/octet-stream'
        
        st.download_button(
            label=f"📥 Download {file_format.upper()}",
            data=data,
            file_name=filename,
            mime=mime_type,
            use_container_width=True
        )
    
    def generate_filename(self, file_format: str, prefix: str = "video_data") -> str:
        """Generate nama file dengan timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = 'csv' if file_format.lower() == 'csv' else 'xlsx'
        return f"{prefix}_{timestamp}.{extension}"
    
    def prepare_summary_statistics(self, video_data: List[Dict]) -> pd.DataFrame:
        """Menyiapkan statistik ringkasan berdasarkan creator dan platform"""
        if not video_data:
            return pd.DataFrame()
        
        # Group by creator
        creator_stats = {}
        for video in video_data:
            creator = video.get('nama_pembuat', 'Unknown')
            platform = video.get('platform', 'Unknown')
            
            if creator not in creator_stats:
                creator_stats[creator] = {
                    'facebook': {'posts': 0, 'views': 0, 'likes': 0, 'comments': 0, 'shares': 0},
                    'zalo': {'posts': 0, 'views': 0, 'likes': 0, 'comments': 0, 'shares': 0},
                    'youtube': {'posts': 0, 'views': 0, 'likes': 0, 'comments': 0, 'shares': 0}
                }
            
            platform_key = platform.lower() if platform.lower() in ['facebook', 'zalo', 'youtube'] else 'facebook'
            creator_stats[creator][platform_key]['posts'] += 1
            creator_stats[creator][platform_key]['views'] += video.get('views', 0)
            creator_stats[creator][platform_key]['likes'] += video.get('likes', 0)
            creator_stats[creator][platform_key]['comments'] += video.get('comments', 0)
            creator_stats[creator][platform_key]['shares'] += video.get('shares', 0)
        
        # Buat DataFrame
        summary_data = []
        for creator, stats in creator_stats.items():
            # Hitung total
            total_posts = sum(stats[platform]['posts'] for platform in stats)
            total_views = sum(stats[platform]['views'] for platform in stats)
            total_likes = sum(stats[platform]['likes'] for platform in stats)
            total_comments = sum(stats[platform]['comments'] for platform in stats)
            total_shares = sum(stats[platform]['shares'] for platform in stats)
            
            row = {
                get_text('summary_creator_name'): creator,
                get_text('summary_fb_post'): stats['facebook']['posts'],
                get_text('summary_fb_view'): f"{stats['facebook']['views']:,}",
                get_text('summary_fb_like'): f"{stats['facebook']['likes']:,}",
                get_text('summary_fb_comment'): f"{stats['facebook']['comments']:,}",
                get_text('summary_fb_share'): f"{stats['facebook']['shares']:,}",
                get_text('summary_zalo_post'): stats['zalo']['posts'],
                get_text('summary_zalo_view'): f"{stats['zalo']['views']:,}",
                get_text('summary_zalo_like'): f"{stats['zalo']['likes']:,}",
                get_text('summary_zalo_comment'): f"{stats['zalo']['comments']:,}",
                get_text('summary_zalo_share'): f"{stats['zalo']['shares']:,}",
                get_text('summary_yt_post'): stats['youtube']['posts'],
                get_text('summary_yt_view'): f"{stats['youtube']['views']:,}",
                get_text('summary_yt_like'): f"{stats['youtube']['likes']:,}",
                get_text('summary_yt_comment'): f"{stats['youtube']['comments']:,}",
                get_text('summary_yt_share'): f"{stats['youtube']['shares']:,}",
                get_text('summary_total_post'): total_posts,
                get_text('summary_total_view'): f"{total_views:,}",
                get_text('summary_total_like'): f"{total_likes:,}",
                get_text('summary_total_comment'): f"{total_comments:,}",
                get_text('summary_total_share'): f"{total_shares:,}"
            }
            summary_data.append(row)
        
        return pd.DataFrame(summary_data)
    
    def export_summary_to_excel(self, video_data: List[Dict]) -> bytes:
        """Ekspor ringkasan statistik ke format Excel sesuai referensi"""
        df = self.prepare_summary_statistics(video_data)
        if df.empty:
            return b''
        
        # Buat workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Ringkasan Statistik"
        
        # Header utama (row 1)
        ws.merge_cells('A1:N1')
        ws['A1'] = get_text('summary_main_header')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Info tanggal (row 2)
        current_date = datetime.now().strftime('%d/%m/%Y')
        ws['A2'] = get_text('summary_date_label')
        ws['B2'] = get_text('summary_time_label')
        ws['C2'] = current_date
        
        # Header platform (row 3)
        ws['A3'] = get_text('summary_creator_name')
        ws['B3'] = 'Facebook'
        ws['F3'] = 'Zalo'
        ws['I3'] = 'Youtube'
        ws['L3'] = get_text('summary_total')
        
        # Merge cells untuk platform headers
        ws.merge_cells('B3:E3')  # Facebook
        ws.merge_cells('F3:H3')  # Zalo
        ws.merge_cells('I3:K3')  # Youtube
        ws.merge_cells('L3:N3')  # Tổng
        
        # Header metrics (row 4)
        metrics_headers = [
            '', 'Post\nPO文', 'Like\n喜歡', 'Comment\n評論', 'Share\n分享',
            'Post\nPO文', 'Like\n喜歡', 'Comment\n評論',
            'Post\nPO文', 'Like\n喜歡', 'Comment\n評論',
            'Post', 'Like', 'Comment'
        ]
        
        for col, header in enumerate(metrics_headers, 1):
            ws.cell(row=4, column=col, value=header)
        
        # Tambahkan data mulai dari row 5
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 5):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Apply styling ke headers
        for row in [3, 4]:
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Simpan ke buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def display_export_summary(self, video_data: List[Dict]) -> None:
        """Menampilkan ringkasan data yang akan diekspor"""
        if not video_data:
            st.warning(get_text('export_no_data'))
            return
        
        total_videos = len(video_data)
        platforms = set(video.get('platform', 'Unknown') for video in video_data)
        date_range = self._get_date_range(video_data)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(get_text('export_total_videos'), total_videos)
        
        with col2:
            st.metric(get_text('export_platforms'), len(platforms))
        
        with col3:
            if date_range:
                st.metric(get_text('export_date_range'), get_text('export_days_format').format(days=date_range['days']))
        
        # Tampilkan platform yang ada
        if platforms:
            st.write(f"**{get_text('export_available_platforms')}:**", ", ".join(sorted(platforms)))
        
        if date_range:
            st.write(f"**{get_text('export_data_period')}:** {date_range['start']} - {date_range['end']}")
    
    def _get_date_range(self, video_data: List[Dict]) -> Dict:
        """Mendapatkan rentang tanggal dari data"""
        dates = []
        for video in video_data:
            waktu = video.get('waktu_penghitungan')
            if waktu:
                if isinstance(waktu, str):
                    try:
                        dates.append(pd.to_datetime(waktu))
                    except:
                        pass
                else:
                    dates.append(waktu)
        
        if not dates:
            return None
        
        min_date = min(dates)
        max_date = max(dates)
        days_diff = (max_date - min_date).days
        
        return {
            'start': min_date.strftime('%Y-%m-%d'),
            'end': max_date.strftime('%Y-%m-%d'),
            'days': days_diff
        }

# Instance global untuk digunakan di seluruh aplikasi
data_exporter = DataExporter()